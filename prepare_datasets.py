"""
Prepare all datasets for ScamStream training experiments.

Converts data from various sources to the simplified JSON format:
  [{"label": "scam"|"harmless", "turns": ["text1", "text2", ...]}, ...]

Outputs (in ScamStream/dataset/):
  - real_1.json         : First half of Real.xlsx (stratified split)
  - real_2.json         : Second half of Real.xlsx (always used as test)
  - tele_data.json      : All Tele-data merged (train + val + test)
  - synthetic_data.json : All Synthetic-data merged (v1_v5 + v2 + v3)
  - real_syn.json       : Real-1 + Synthetic-data
  - real_tele.json      : Real-1 + Tele-data

Usage:
    python prepare_datasets.py
"""

import glob
import json
import os
import random
import re
import unicodedata
from collections import Counter

# ── Paths ────────────────────────────────────────────────────────────

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
BASELINE2    = os.path.dirname(SCRIPT_DIR)
EXCEL_PATH   = os.path.join(BASELINE2, "Real.xlsx")
TELE_DIR     = os.path.join(BASELINE2, "Tele-data")
SYNTH_DIR    = os.path.join(BASELINE2, "Synthetic-data")
OUT_DIR      = os.path.join(SCRIPT_DIR, "dataset")

SEED = 42


# ── Text utilities ───────────────────────────────────────────────────

def clean(text: str) -> str:
    """Normalize & clean a single turn text."""
    text = unicodedata.normalize("NFC", str(text))
    text = re.sub(r"[\x00-\x09\x0b-\x0c\x0e-\x1f\x7f]", "", text)
    return re.sub(r"\s+", " ", text).strip()


# ── 1. Real.xlsx → JSON ──────────────────────────────────────────────

def load_excel(path: str) -> list:
    """Read Real.xlsx (Sheet1=scam, no_scam=harmless) → simplified format."""
    try:
        import openpyxl
    except ImportError:
        print("  [ERROR] openpyxl not installed. Run: pip install openpyxl")
        raise

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = []
    sheet_label = {"Sheet1": "scam", "no_scam": "harmless"}

    for sheet_name, label in sheet_label.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet '{sheet_name}' not found, skipping.")
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, values_only=True):
            conv_text = row[1] if len(row) > 1 else row[0]
            if not conv_text or not str(conv_text).strip():
                continue
            raw_turns = [t.strip() for t in str(conv_text).strip().split("\n") if t.strip()]
            turns = [clean(t) for t in raw_turns if clean(t)]
            if turns:
                records.append({"label": label, "turns": turns})

    return records


# ── 2. Stratified 50/50 split ────────────────────────────────────────

def stratified_split(data: list, seed: int = 42):
    """Split data into two halves, maintaining scam/harmless ratio."""
    rng = random.Random(seed)

    scam     = [d for d in data if d["label"] == "scam"]
    harmless = [d for d in data if d["label"] == "harmless"]

    rng.shuffle(scam)
    rng.shuffle(harmless)

    mid_scam = len(scam) // 2
    mid_harm = len(harmless) // 2

    half_1 = scam[:mid_scam] + harmless[:mid_harm]
    half_2 = scam[mid_scam:] + harmless[mid_harm:]

    rng.shuffle(half_1)
    rng.shuffle(half_2)

    return half_1, half_2


# ── 3. Tele-data → JSON ─────────────────────────────────────────────

def load_tele_data(tele_dir: str) -> list:
    """
    Read train.json, val.json, test.json from Tele-data.
    These have rich format with turns[].content and label field.
    Convert to simplified format.
    """
    records = []
    for fname in ["train.json", "val.json", "test.json"]:
        fpath = os.path.join(tele_dir, fname)
        if not os.path.exists(fpath):
            print(f"  [WARN] {fpath} not found, skipping.")
            continue
        with open(fpath, encoding="utf-8") as f:
            data = json.load(f)
        for d in data:
            label = d.get("label", "")
            if label not in ("scam", "harmless"):
                continue
            turns = []
            for t in d.get("turns", []):
                text = t.get("content", "").strip()
                if text:
                    turns.append(clean(text))
            if turns:
                records.append({"label": label, "turns": turns})
        print(f"  Loaded {fname}: {len(data)} dialogues")

    return records


# ── 4. Synthetic-data → JSON ─────────────────────────────────────────

def load_synthetic_data(synth_dir: str) -> list:
    """
    Read all conversation folders from all subfolders in Synthetic-data.
    Each conversation has history[].text and caller_card.is_scam.
    """
    records = []
    # Process all subdirectories (v1_v5, v2, v3)
    for sub in sorted(os.listdir(synth_dir)):
        sub_path = os.path.join(synth_dir, sub)
        if not os.path.isdir(sub_path):
            continue

        files = glob.glob(os.path.join(sub_path, "*", "*.json"))
        count = 0
        for fpath in files:
            # Skip log.txt etc
            if not fpath.endswith(".json"):
                continue
            try:
                with open(fpath, encoding="utf-8") as f:
                    d = json.load(f)
            except (json.JSONDecodeError, UnicodeDecodeError):
                continue

            is_scam = d.get("caller_card", {}).get("is_scam", None)
            if is_scam is None:
                continue
            label = "scam" if is_scam else "harmless"
            turns = []
            for t in d.get("history", []):
                text = t.get("text", "").strip()
                if text:
                    turns.append(clean(text))
            if turns:
                records.append({"label": label, "turns": turns})
                count += 1
        print(f"  {sub}: {count} dialogues")

    return records


# ── 5. Save utility ─────────────────────────────────────────────────

def save_json(data: list, path: str):
    """Save data to JSON file."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def print_stats(name: str, data: list):
    """Print label distribution stats."""
    dist = Counter(d["label"] for d in data)
    total = len(data)
    print(f"  {name}: {total} dialogues (scam={dist.get('scam', 0)}, harmless={dist.get('harmless', 0)})")


# ── Main ─────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    print("=" * 60)
    print("PREPARING DATASETS FOR SCAMSTREAM")
    print("=" * 60)

    # ── Step 1: Convert Real.xlsx ──
    print("\n[1/5] Converting Real.xlsx...")
    real_data = load_excel(EXCEL_PATH)
    print_stats("Real.xlsx total", real_data)

    # ── Step 2: Stratified split → Real-1, Real-2 ──
    print("\n[2/5] Splitting Real data into Real-1 and Real-2 (stratified 50/50)...")
    real_1, real_2 = stratified_split(real_data, seed=SEED)
    print_stats("Real-1 (train)", real_1)
    print_stats("Real-2 (test)", real_2)

    save_json(real_1, os.path.join(OUT_DIR, "real_1.json"))
    save_json(real_2, os.path.join(OUT_DIR, "real_2.json"))
    print("  ✓ Saved real_1.json, real_2.json")

    # ── Step 3: Convert Tele-data ──
    print("\n[3/5] Converting Tele-data...")
    tele_data = load_tele_data(TELE_DIR)
    print_stats("Tele-data total", tele_data)

    save_json(tele_data, os.path.join(OUT_DIR, "tele_data.json"))
    print("  ✓ Saved tele_data.json")

    # ── Step 4: Convert Synthetic-data ──
    print("\n[4/5] Converting Synthetic-data...")
    synth_data = load_synthetic_data(SYNTH_DIR)
    print_stats("Synthetic-data total", synth_data)

    save_json(synth_data, os.path.join(OUT_DIR, "synthetic_data.json"))
    print("  ✓ Saved synthetic_data.json")

    # ── Step 5: Create combined datasets ──
    print("\n[5/5] Creating combined datasets...")

    real_syn = real_1 + synth_data
    print_stats("Real-1 + Synthetic", real_syn)
    save_json(real_syn, os.path.join(OUT_DIR, "real_syn.json"))
    print("  ✓ Saved real_syn.json")

    real_tele = real_1 + tele_data
    print_stats("Real-1 + Tele", real_tele)
    save_json(real_tele, os.path.join(OUT_DIR, "real_tele.json"))
    print("  ✓ Saved real_tele.json")

    # ── Summary ──
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"\nOutput directory: {OUT_DIR}\n")
    files = sorted(os.listdir(OUT_DIR))
    for fname in files:
        fpath = os.path.join(OUT_DIR, fname)
        size_mb = os.path.getsize(fpath) / (1024 * 1024)
        with open(fpath, encoding="utf-8") as f:
            data = json.load(f)
        dist = Counter(d["label"] for d in data)
        print(f"  {fname:25s}  {len(data):5d} dlg  "
              f"(scam={dist.get('scam',0):4d}, harmless={dist.get('harmless',0):4d})  "
              f"{size_mb:.1f} MB")

    print("\nDone! ✓")


if __name__ == "__main__":
    main()
