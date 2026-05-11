"""
Test M1 Multi-Class Classifier trên dữ liệu test.

Usage:
    python test.py --data path/to/test.json --model outputs/best_model
    python test.py --data path/to/test.json --model outputs/best_model --out-dir outputs/eval
"""

import argparse
import dataclasses
import json
import os
import sys

import numpy as np
import torch
from torch.utils.data import DataLoader
from transformers import AutoTokenizer

_dir = os.path.dirname(os.path.abspath(__file__))
if _dir not in sys.path:
    sys.path.insert(0, _dir)

from config import M1Config, CLASS_NAMES, SCENARIO_TO_IDX
from dataset import DialogueDataset, collate_fn, load_json, get_class_label
from model import M1Classifier
from metrics import compute_streaming_metrics, print_streaming_report


@torch.no_grad()
def run_inference(model, dataset, raw, device, cfg):
    model.eval()
    loader = DataLoader(dataset, batch_size=2, shuffle=False,
                        collate_fn=collate_fn, num_workers=0)

    all_labels, all_preds, all_probs, all_t_probs, all_p_agg = [], [], [], [], []
    total_loss, n_dlg = 0.0, 0

    for batch in loader:
        input_ids  = batch["input_ids"].to(device)
        attn_masks = batch["attn_masks"].to(device)
        turn_mask  = batch["turn_mask"].to(device)
        labels     = batch["labels"].to(device)
        n_turns    = batch["n_turns"]

        output = model(input_ids, attn_masks, turn_mask, labels=labels)
        B = labels.shape[0]
        if output["loss"] is not None:
            total_loss += output["loss"].item() * B
        n_dlg += B

        for b in range(B):
            n = int(n_turns[b].item())
            all_labels.append(int(labels[b].item()))
            all_preds.append(int(output["dialogue_preds"][b].item()))
            all_probs.append(output["dialogue_probs"][b].cpu().numpy())
            t_probs = output["turn_probs"][b, :n].cpu().numpy()    # [T, C]
            p_agg   = output["p_agg"][b, :n].cpu().numpy()          # [T, C]
            all_t_probs.append(t_probs)
            all_p_agg.append(p_agg)

    metrics = compute_streaming_metrics(
        all_labels, all_preds, all_probs, all_p_agg,
        class_names=cfg.class_names,
        scam_alert_thresh=cfg.scam_alert_thresh,
    )
    metrics["loss"] = total_loss / max(n_dlg, 1)
    return metrics, all_labels, all_preds, all_probs, all_t_probs, all_p_agg


def save_metrics_json(metrics, out_path):
    summary = {
        "accuracy":              round(metrics["accuracy"], 4),
        "macro_f1":              round(metrics["macro_f1"], 4),
        "weighted_f1":           round(metrics["weighted_f1"], 4),
        "overall_detection_rate": round(metrics["overall_detection_rate"], 4),
        "false_alarm_rate":      round(metrics["false_alarm_rate"], 4),
        "loss":                  round(metrics.get("loss", float("nan")), 4),
        "per_class_f1":          metrics["per_class_f1"],
        "per_class_precision":   metrics["per_class_precision"],
        "per_class_recall":      metrics["per_class_recall"],
        "per_class_support":     metrics["per_class_support"],
        "confusion_matrix":      metrics["confusion_matrix"],
    }
    if "scam_detection" in metrics:
        summary["scam_detection"] = metrics["scam_detection"]
    if "mean_alert_turn" in metrics:
        summary["mean_alert_turn"]   = round(metrics["mean_alert_turn"], 4)
        summary["median_alert_turn"] = round(metrics["median_alert_turn"], 4)
        summary["alert_at_half"]     = round(metrics["alert_at_half"], 4)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2, default=str)
    print(f"  Metrics saved → {out_path}")


def save_errors_excel(raw, all_labels, all_preds, all_t_probs, all_p_agg,
                      cfg, out_path):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [WARN] openpyxl not installed — skipping Excel output")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wrong Predictions"

    # Header
    class_headers = [f"p_agg_{c}" for c in cfg.class_names]
    headers = ["#", "Error Type", "True Class", "Pred Class",
               "Turn #"] + class_headers + ["Turn Text"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row_idx = 2
    err_count = 0
    for idx, (label, pred, t_probs, p_agg) in enumerate(
            zip(all_labels, all_preds, all_t_probs, all_p_agg)):
        if pred == label:
            continue

        err_count += 1
        dlg = raw[idx]
        turns = dlg["turns"][:t_probs.shape[0]]
        true_name = cfg.class_names[label]
        pred_name = cfg.class_names[pred]
        err_type = f"{true_name} → {pred_name}"

        for t, turn_text in enumerate(turns):
            is_first = (t == 0)
            # Normalize p_agg at this turn
            p_row = p_agg[t]
            p_norm = p_row / (p_row.sum() + 1e-8)

            row_data = [
                err_count if is_first else "",
                err_type  if is_first else "",
                true_name if is_first else "",
                pred_name if is_first else "",
                t + 1,
            ] + [round(float(p_norm[c]), 4) for c in range(cfg.num_classes)] + [
                str(turn_text),
            ]
            ws.append(row_data)
            row_idx += 1

        # Blank separator row
        ws.append([""] * len(headers))
        row_idx += 1

    # Column widths
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    ws2.append(["Total wrong", err_count])

    wb.save(out_path)
    print(f"  Errors saved  → {out_path}  ({err_count} wrong predictions)")


def print_error_samples(raw, all_labels, all_preds, all_t_probs, all_p_agg,
                        cfg, max_per_type=10):
    """Print error samples grouped by error type."""
    from collections import defaultdict
    errors = defaultdict(list)

    for idx, (label, pred, t_probs, p_agg) in enumerate(
            zip(all_labels, all_preds, all_t_probs, all_p_agg)):
        if pred == label:
            continue
        true_name = cfg.class_names[label]
        pred_name = cfg.class_names[pred]
        errors[f"{true_name}→{pred_name}"].append(
            (idx, t_probs, p_agg, raw[idx]["turns"][:t_probs.shape[0]])
        )

    print(f"\n  Total errors: {sum(len(v) for v in errors.values())}")
    for err_type, samples in sorted(errors.items()):
        print(f"\n{'='*65}")
        print(f"  {err_type}: {len(samples)} errors (showing max {max_per_type})")
        print(f"{'='*65}")
        for idx, tps, pag, turns in samples[:max_per_type]:
            n = len(turns)
            # Final prediction
            p_final = pag[-1]
            p_norm = p_final / (p_final.sum() + 1e-8)
            probs_str = " ".join(f"{cfg.class_names[c]}={p_norm[c]:.3f}" for c in range(cfg.num_classes))
            print(f"\n  [#{idx}] {probs_str}  ({n} turns)")
            for t, turn in enumerate(turns[:8], 1):  # max 8 turns shown
                p_row = pag[min(t-1, pag.shape[0]-1)]
                p_n = p_row / (p_row.sum() + 1e-8)
                top_c = int(p_n.argmax())
                print(f"    T{t}: → {cfg.class_names[top_c]}({p_n[top_c]:.3f})  {str(turn)[:80]}")


def parse_args():
    cfg = M1Config()
    parser = argparse.ArgumentParser(description="Test M1 Multi-Class Scam Classifier")
    parser.add_argument("--data",      required=True,
                        help="Path to test JSON file")
    parser.add_argument("--model",     default=os.path.join(cfg.output_dir, "best_model"),
                        help="Path to model directory (default: outputs/best_model)")
    parser.add_argument("--out-dir",   default=None,
                        help="Output directory (default: same as --model)")
    return parser.parse_args()


def main():
    args = parse_args()
    out_dir = args.out_dir or args.model

    print("=" * 65)
    print("M1 MULTI-CLASS SCAM CLASSIFIER — TEST")
    print("=" * 65)
    print(f"  Data:      {args.data}")
    print(f"  Model:     {args.model}")
    print(f"  Out dir:   {out_dir}")

    if not os.path.exists(args.data):
        print(f"\n  [ERROR] Not found: {args.data}")
        sys.exit(1)

    model_pt = os.path.join(args.model, "model.pt")
    if not os.path.exists(model_pt):
        print(f"\n  [ERROR] Model not found: {model_pt}")
        sys.exit(1)

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"  Device:    {device}")

    config_json = os.path.join(args.model, "config.json")
    if os.path.exists(config_json):
        with open(config_json) as f:
            cfg_dict = json.load(f)
        cfg = M1Config(**{k: v for k, v in cfg_dict.items()
                          if k in {f.name for f in dataclasses.fields(M1Config)}})
    else:
        cfg = M1Config()

    tokenizer = AutoTokenizer.from_pretrained(args.model)
    model     = M1Classifier(cfg).to(device)
    model.load_state_dict(torch.load(model_pt, map_location=device, weights_only=True))
    model.eval()

    raw = load_json(args.data)
    from collections import Counter
    class_dist = Counter(get_class_label(d) for d in raw)
    print(f"\n  {len(raw)} conversations:")
    for c_idx, c_name in enumerate(cfg.class_names):
        print(f"    {c_name}: {class_dist.get(c_idx, 0)}")

    dataset = DialogueDataset(raw, tokenizer, cfg.max_turn_len, cfg.max_turns)

    print(f"\nEvaluating...")
    metrics, all_labels, all_preds, all_probs, all_t_probs, all_p_agg = run_inference(
        model, dataset, raw, device, cfg
    )

    print_streaming_report(metrics)

    os.makedirs(out_dir, exist_ok=True)
    save_metrics_json(metrics, os.path.join(out_dir, "test_metrics.json"))
    save_errors_excel(raw, all_labels, all_preds, all_t_probs, all_p_agg,
                      cfg, os.path.join(out_dir, "test_errors.xlsx"))

    # Error analysis
    print_error_samples(raw, all_labels, all_preds, all_t_probs, all_p_agg, cfg)

    print("\nDone!")


if __name__ == "__main__":
    main()
