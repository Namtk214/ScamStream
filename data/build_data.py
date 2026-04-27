"""
Build data_processed/train.json và test.json với schema đơn giản:
  [{"label": "scam"|"harmless", "turns": ["text1", "text2", ...]}, ...]

train.json = train.json cũ + data_viscamdial_1000_v1/
test.json  = Excel (Sheet1 = scam, no_scam = harmless)
"""

import glob
import json
import os
import re
import unicodedata

DATA_DIR   = os.path.dirname(os.path.abspath(__file__))
OUT_DIR    = os.path.join(DATA_DIR, "data_processed")
TRAIN_JSON = os.path.join(DATA_DIR, "train.json")
VISCAM_DIR = os.path.join(DATA_DIR, "data_viscamdial_1000_v1")
EXCEL_PATH = os.path.join(DATA_DIR, "Tổng hợp kịch bản test AI on devices_result_v2.xlsx")


def clean(text: str) -> str:
    text = unicodedata.normalize("NFC", str(text))
    text = re.sub(r"[\x00-\x09\x0b-\x0c\x0e-\x1f\x7f]", "", text)
    return re.sub(r"\s+", " ", text).strip()


# ── 1. Đọc train.json cũ ─────────────────────────────────────────

def load_old_train(path: str) -> list:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    records = []
    for d in data:
        label = d.get("label", "")
        if label not in ("scam", "harmless"):
            continue
        turns = [clean(t["content"]) for t in d["turns"] if t.get("content", "").strip()]
        if turns:
            records.append({"label": label, "turns": turns})
    return records


# ── 2. Đọc viscamdial folder ─────────────────────────────────────

def load_viscamdial(folder: str) -> list:
    files = glob.glob(os.path.join(folder, "*", "*.json"))
    records = []
    for path in files:
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
        is_scam = d.get("caller_card", {}).get("is_scam", None)
        if is_scam is None:
            continue
        label = "scam" if is_scam else "harmless"
        turns = [clean(t["text"]) for t in d.get("history", []) if t.get("text", "").strip()]
        if turns:
            records.append({"label": label, "turns": turns})
    return records


# ── 3. Đọc Excel → test set ──────────────────────────────────────

def load_excel(path: str) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = []
    sheet_label = {"Sheet1": "scam", "no_scam": "harmless"}
    for sheet_name, label in sheet_label.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Không có sheet '{sheet_name}', bỏ qua.")
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, values_only=True):
            conv_text = row[1] if len(row) > 1 else row[0]
            if not conv_text or not str(conv_text).strip():
                continue
            raw_turns = [t.strip() for t in str(conv_text).strip().split("\n") if t.strip()]
            turns = [clean(t) for t in raw_turns if clean(t)]
            if turns:
                records.append({"label": label, "turns": turns})
    return records


# ── Main ─────────────────────────────────────────────────────────

def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    print("=== Building train.json ===")
    old_train = load_old_train(TRAIN_JSON)
    print(f"  train.json cũ: {len(old_train)} dialogues")

    viscam = load_viscamdial(VISCAM_DIR)
    print(f"  viscamdial:    {len(viscam)} dialogues")

    train = old_train + viscam
    from collections import Counter
    dist = Counter(d["label"] for d in train)
    print(f"  Total train:   {len(train)} (scam={dist['scam']}, harmless={dist['harmless']})")

    out_train = os.path.join(OUT_DIR, "train.json")
    with open(out_train, "w", encoding="utf-8") as f:
        json.dump(train, f, ensure_ascii=False, indent=2)
    print(f"  Saved → {out_train}")

    print("\n=== Building test.json ===")
    test = load_excel(EXCEL_PATH)
    dist_test = Counter(d["label"] for d in test)
    print(f"  Total test: {len(test)} (scam={dist_test['scam']}, harmless={dist_test['harmless']})")

    out_test = os.path.join(OUT_DIR, "test.json")
    with open(out_test, "w", encoding="utf-8") as f:
        json.dump(test, f, ensure_ascii=False, indent=2)
    print(f"  Saved → {out_test}")

    print("\n=== Sample ===")
    s = train[0]
    print(f"  label: {s['label']}, turns: {len(s['turns'])}")
    print(f"  turn[0]: {s['turns'][0][:80]}")
    s2 = test[0]
    print(f"\n  [test] label: {s2['label']}, turns: {len(s2['turns'])}")
    print(f"  turn[0]: {s2['turns'][0][:80]}")


if __name__ == "__main__":
    main()
